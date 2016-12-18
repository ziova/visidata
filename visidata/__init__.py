#!/usr/bin/env python3
#
# VisiData: a curses interface for exploring and arranging tabular data
#
# Copyright (C) 2016 Saul Pwanson
#
#   This program is free software: you can redistribute it and/or modify
#   it under the terms of the GNU General Public License as published by
#   the Free Software Foundation, either version 3 of the License, or
#   (at your option) any later version.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License
#   along with this program.  If not, see <http://www.gnu.org/licenses/>.

__version__ = '0.39'
__author__ = 'Saul Pwanson <vd@saul.pw>'
__license__ = 'GPLv3'
__status__ = 'Development'

import itertools
import os.path
from copy import copy
import io
import string
import collections
import functools
import statistics
import curses

import re
import csv

class attrdict(object):
    def __init__(self, d):
        self.__dict__ = d

base_options = collections.OrderedDict()
options = attrdict(base_options)

anytype = lambda r=None: str(r)
anytype.__name__ = ''

class WrongTypeStr(str):
    pass

class CalcErrorStr(str):
    pass


# VisiData singleton contains all sheets
@functools.lru_cache()
def vd():
    return VisiData()


def status(s):
    return vd().status(s)


from visidata.tui import edit_text, Key, Shift, Ctrl, keyname, EscapeException, wrapper, colors, draw_clip
from visidata.date import date
from .Sheet import Sheet, base_commands
from .Column import Column, ColumnAttr, AttrColumns, PyobjColumns, ArrayColumns, ArrayNamedColumns


initialStatus = 'saul.pw/VisiData v' + __version__
defaultColNames = list(itertools.chain(string.ascii_uppercase, [''.join(i) for i in itertools.product(string.ascii_uppercase, repeat=2)]))

scr = None   # toplevel curses screen

windowWidth = None
windowHeight = None


class CommandPrefixException(Exception):
    pass

def error(s):
    raise Exception(s)

def setup_sheet_commands():
    def command(ch, cmdstr, helpstr):
        base_commands[('', ch)] = (cmdstr, helpstr)
    def global_command(ch, cmdstr, helpstr):
        base_commands[('g', ch)] = (cmdstr, helpstr)
    def option(name, default, helpstr=''):
        base_options[name] = default
#        base_options[name] = (default, helpstr)
    def theme(name, default, helpstr=''):
        base_options[name] = default
#        base_options[name] = (default, helpstr)

    option('csv_dialect', 'excel', 'dialect passed to csv.reader')
    option('csv_delimiter', ',', 'delimiter passed to csv.reader')
    option('csv_quotechar', '"', 'quotechar passed to csv.reader')
    option('csv_header', '', 'parse first row of CSV as column names')

    option('debug', '', 'abort on error and display stacktrace')
    option('readonly', '', 'disable saving')

    option('encoding', 'utf-8', 'as passed to codecs.open')
    option('encoding_errors', 'surrogateescape', 'as passed to codecs.open')

    option('ColumnStats', False, 'include mean/median/etc on Column sheet')

    option('SubsheetSep', '~')

    # display/color scheme
    theme('SheetNameFmt', '%s| ', 'status line prefix')
    theme('ch_VisibleNone', '',  'visible contents of a cell whose value was None')
    theme('ch_FunctionError', '¿', 'when computation fails due to exception')
    theme('ch_Histogram', '*')
    theme('ch_ColumnFiller', ' ', 'pad chars after column value')
    theme('ch_LeftMore', '<')
    theme('ch_RightMore', '>')
    theme('ch_ColumnSep', '|', 'chars between columns')
    theme('ch_Ellipsis', '…')
    theme('ch_StatusSep', ' | ')
    theme('ch_KeySep', '/')
    theme('ch_EditPadChar', '_')
    theme('ch_Newline', '\\n')
    theme('ch_Unprintable', '.')
    theme('ch_WrongType', '~')
    theme('ch_Error', '!')

    theme('c_default', 'normal')
    theme('c_Header', 'bold')
    theme('c_CurHdr', 'reverse')
    theme('c_CurRow', 'reverse')
    theme('c_CurCol', 'bold')
    theme('c_KeyCols', 'brown')
    theme('c_StatusLine', 'bold')
    theme('c_SelectedRow', 'green')
    theme('c_ColumnSep', 'blue')
    theme('c_EditCell', 'normal')
    theme('c_WrongType', 'magenta')
    theme('c_Error', 'red')

    command(Key.F1,    'vd.push(CommandHelp(sheet))', 'open help sheet')
    command(Key('q'),  'vd.sheets.pop(0)', 'drop this sheet')

    command(Key.LEFT,  'cursorRight(-1)', 'go one column left')
    command(Key.DOWN,  'cursorDown(+1)', 'go one row down')
    command(Key.UP,    'cursorDown(-1)', 'go one row up')
    command(Key.RIGHT, 'cursorRight(+1)', 'go one column right')
    command(Key.NPAGE, 'cursorDown(nVisibleRows); topRowIndex += nVisibleRows', 'scroll one page down')
    command(Key.PPAGE, 'cursorDown(-nVisibleRows); topRowIndex -= nVisibleRows', 'scroll one page up')
    command(Key.HOME,  'topRowIndex = cursorRowIndex = 0', 'go to top row')
    command(Key.END,   'cursorRowIndex = len(rows)-1', 'go to last row')

    command(Key('h'), 'cursorRight(-1)', 'go one column left')
    command(Key('j'), 'cursorDown(+1)', 'go one row down')
    command(Key('k'), 'cursorDown(-1)', 'go one row up')
    command(Key('l'), 'cursorRight(+1)', 'go one column right')

    command(Shift.H, 'moveVisibleCol(cursorVisibleColIndex, max(cursorVisibleColIndex-1, 0)); cursorVisibleColIndex -= 1', 'move this column one left')
    command(Shift.J, 'cursorRowIndex = moveListItem(rows, cursorRowIndex, min(cursorRowIndex+1, nRows-1))', 'move this row one down')
    command(Shift.K, 'cursorRowIndex = moveListItem(rows, cursorRowIndex, max(cursorRowIndex-1, 0))', 'move this row one up')
    command(Shift.L, 'moveVisibleCol(cursorVisibleColIndex, min(cursorVisibleColIndex+1, nVisibleCols-1)); cursorVisibleColIndex += 1', 'move this column one right')

    command(Ctrl.G, 'status(statusLine)', 'show this sheet info')
    command(Ctrl.P, 'status(vd.statusHistory[0])', 'show previous status line again')
    command(Ctrl.V, 'status(initialStatus)', 'show version information')

    command(Key('t'), 'topRowIndex = cursorRowIndex', 'scroll cursor row to top of screen')
    command(Key('m'), 'topRowIndex = cursorRowIndex-int(nVisibleRows/2)', 'scroll cursor row to middle of screen')
    command(Key('b'), 'topRowIndex = cursorRowIndex-nVisibleRows+1', 'scroll cursor row to bottom of screen')

    command(Key('<'), 'skipUp()', 'skip up this column to previous value')
    command(Key('>'), 'skipDown()', 'skip down this column to next value')

    command(Key('_'), 'cursorCol.width = cursorCol.getMaxWidth(visibleRows)', 'set this column width to fit visible cells')
    command(Key('-'), 'cursorCol.width = 0', 'hide this column')
    command(Key('^'), 'cursorCol.name = cursorCol.getDisplayValue(cursorRow)', 'set this column header to this cell value')
    command(Key('!'), 'toggleKeyColumn(cursorColIndex)', 'toggle this column as a key column')

    command(Key('@'), 'cursorCol.type = date', 'set column type to ISO8601 datetime')
    command(Key('#'), 'cursorCol.type = int', 'set column type to integer')
    command(Key('$'), 'cursorCol.type = str', 'set column type to string')
    command(Key('%'), 'cursorCol.type = float', 'set column type to float')
    command(Key('~'), 'cursorCol.type = detectType(cursorValue)', 'autodetect type of column by its data')

    command(Key('['), 'rows = sorted(rows, key=lambda r: cursorCol.getValue(r))', 'sort by this column ascending')
    command(Key(']'), 'rows = sorted(rows, key=lambda r: cursorCol.getValue(r), reverse=True)', 'sort by this column descending')
    command(Ctrl.E, 'options.debug = True; error(vd.lastErrors[-1])', 'abort and print last error to terminal')
    command(Ctrl.D, 'options.debug = not options.debug; status("debug " + ("ON" if options.debug else "OFF"))', 'toggle debug mode')

    command(Shift.E, 'if vd.lastErrors: vd.push(SheetText("last_error", vd.lastErrors[-1]))', 'open stack trace for most recent error')
    command(Shift.F, 'vd.push(SheetFreqTable(sheet, cursorCol))', 'open frequency table from values in this column')

    command(Key('d'), 'rows.pop(cursorRowIndex)', 'delete this row')

    command(Key('g'), 'raise CommandPrefixException("g")', 'add global prefix')

    command(Shift.S, 'vd.push(Sheets(vd.sheets))', 'open Sheet stack')
    command(Shift.C, 'vd.push(SheetColumns(sheet))', 'open Columns for this sheet')
    command(Shift.O, 'vd.push(SheetDict("options", options.__dict__))', 'open Options')

    command(Key('/'), 'searchRegex(inputLine(prompt="/"), columns=[cursorCol], moveCursor=True)', 'search this column forward for regex')
    command(Key('?'), 'searchRegex(inputLine(prompt="?"), columns=[cursorCol], backward=True, moveCursor=True)', 'search this column backward for regex')
    command(Key('n'), 'searchRegex(columns=[cursorCol], moveCursor=True)', 'go to next match')
    command(Key('p'), 'searchRegex(columns=[cursorCol], backward=True, moveCursor=True)', 'go to previous match')

    command(Key(' '), 'toggle([cursorRow]); cursorDown(1)', 'toggle select of this row')
    command(Key('s'), 'select([cursorRow]); cursorDown(1)', 'select this row')
    command(Key('u'), 'unselect([cursorRow]); cursorDown(1)', 'unselect this row')
    command(Key('|'), 'select(rows[r] for r in searchRegex(inputLine(prompt="|"), columns=[cursorCol]))', 'select rows by regex in this column')
    command(Key('\\'), 'unselect(rows[r] for r in searchRegex(inputLine(prompt="\\\\"), columns=[cursorCol]))', 'unselect rows by regex in this column')

    command(Shift.R, 'source.type = inputLine("change type to: ", value=initialsource.type)', 'set source type of this sheet')
    command(Ctrl.R, 'open_source(vd.sheets.pop(0).source); status("reloaded")', 'reload sheet from source')
    command(Ctrl.S, 'saveSheet(sheet, inputLine("save to: "))', 'save this sheet to new file')
    command(Key('o'), 'open_source(inputLine("open: "))', 'open local file or url')
    command(Ctrl.O, 'expr = inputLine("eval: "); pushPyObjSheet(expr, eval(expr))', 'eval Python expression and open the result')

    command(Key('e'), 'cursorCol.setValue(cursorRow, editCell(cursorVisibleColIndex))', 'edit this cell')
    command(Key('c'), 'cursorVisibleColIndex = findColIdx(inputLine("goto column name: "), visibleCols)', 'goto visible column by name')
    command(Key('r'), 'cursorRowIndex = int(inputLine("goto row number: "))', 'goto row number')

    command(Key('='), 'addColumn(ColumnExpr(sheet, inputLine("new column expr=")), index=cursorColIndex+1)', 'add column by expr')
    command(Key(':'), 'addColumn(ColumnRegex(sheet, inputLine("new column regex:")), index=cursorColIndex+1)', 'add column by regex')
    command(Ctrl('^'), 'vd.sheets[0], vd.sheets[1] = vd.sheets[1], vd.sheets[0]', 'jump to previous sheet')
    command(Key.TAB,  'moveListItem(vd.sheets, 0, len(vd.sheets))', 'cycle through sheet stack')
    command(Key.BTAB, 'moveListItem(vd.sheets, -1, 0)', 'reverse cycle through sheet stack')

# when used with 'g' prefix
    global_command(Key('q'), 'vd.sheets.clear()', 'drop all sheets (clean exit)')

    global_command(Key('h'), 'cursorVisibleColIndex = leftVisibleColIndex = 0', 'go to leftmost column')
    global_command(Key('k'), 'cursorRowIndex = topRowIndex = 0', 'go to top row')
    global_command(Key('j'), 'cursorRowIndex = len(rows); topRowIndex = cursorRowIndex-nVisibleRows', 'go to bottom row')
    global_command(Key('l'), 'cursorVisibleColIndex = len(visibleCols)-1', 'go to rightmost column')

    global_command(Shift.H, 'moveListItem(columns, cursorColIndex, 0)', 'move this column all the way to the left')
    global_command(Shift.J, 'moveListItem(rows, cursorRowIndex, nRows)', 'move this row all the way to the bottom')
    global_command(Shift.K, 'moveListItem(rows, cursorRowIndex, 0)', 'move this row all the way to the top')
    global_command(Shift.L, 'moveListItem(columns, cursorColIndex, nCols)', 'move this column all the way to the right')

    global_command(Key('_'), 'for c in visibleCols: c.width = c.getMaxWidth(visibleRows)', 'set width of all columns to fit visible cells')
    global_command(Key('^'), 'for c in visibleCols: c.name = c.getDisplayValue(cursorRow)', 'set names of all visible columns to this row')
    global_command(Key('~'), 'for c in visibleCols: c.type = detectType(c.getValue(cursorRow))', 'autodetect types of all visible columns by their data')

    global_command(Shift.E, 'vd.push(SheetText("last_error", "\\n\\n".join(vd.lastErrors)))', 'open last 10 errors')

    global_command(Key('/'), 'searchRegex(inputLine(prompt="/"), moveCursor=True, columns=visibleCols)', 'search regex forward in all visible columns')
    global_command(Key('?'), 'searchRegex(inputLine(prompt="?"), backward=True, moveCursor=True, columns=visibleCols)', 'search regex backward in all visible columns')
    global_command(Key('n'), 'cursorRowIndex = max(searchRegex())', 'go to first match')
    global_command(Key('p'), 'cursorRowIndex = min(searchRegex())', 'go to last match')

    global_command(Key(' '), 'toggle(rows)', 'toggle select of all rows')
    global_command(Key('s'), 'select(rows)', 'select all rows')
    global_command(Key('u'), '_selectedRows = {}', 'unselect all rows')

    global_command(Key('|'), 'select(rows[r] for r in searchRegex(inputLine(prompt="|"), columns=visibleCols))', 'select rows by regex in all visible columns')
    global_command(Key('\\'), 'unselect(rows[r] for r in searchRegex(inputLine(prompt="\\\\"), columns=visibleCols))', 'unselect rows by regex in all visible columns')

    global_command(Key('d'), 'rows = [r for r in rows if not isSelected(r)]; _selectedRows = {}', 'delete all selected rows')

    global_command(Ctrl.P, 'vd.push(SheetText("statuses", vd.statusHistory))', 'open last 100 statuses')

    # experimental commands
    command(Key('"'), 'rows.insert(cursorRowIndex, copy(cursorRow))', 'insert duplicate of this row')

# end setup_sheet_commands

### VisiData core

def moveListItem(L, fromidx, toidx):
    r = L.pop(fromidx)
    L.insert(toidx, r)
    return toidx

def detectType(v):
    def tryType(T, v):
        try:
            v = T(v)
            return T
        except:
            return None

    return tryType(int, v) or tryType(float, v) or tryType(date, v) or str


class VisiData:
    def __init__(self):
        self.sheets = []
        self.statusHistory = []
        self._status = []
        self.status(initialStatus)
        self.lastErrors = []

    def status(self, s):
        strs = str(s)
        self._status.append(strs)
        self.statusHistory.insert(0, strs)
        del self.statusHistory[100:]  # keep most recent 100 only
        return s

    def editText(self, y, x, w, **kwargs):
        v = editText(scr, y, x, w, **kwargs)
        self.status('"%s"' % v)
        return v

    def exceptionCaught(self, status=True):
        import traceback
        self.lastErrors.append(traceback.format_exc().strip())
        self.lastErrors = self.lastErrors[-10:]  # keep most recent
        if status:
            self.status(self.lastErrors[-1].splitlines()[-1])
        if options.debug:
            raise

    def drawLeftStatus(self, scr, sheet):
        'draws sheet info on last line, including previous status messages, which are then cleared.'
        attr = colors[options.c_StatusLine]
        statusstr = options.SheetNameFmt % sheet.name + options.ch_StatusSep.join(self._status)
        draw_clip(scr, windowHeight-1, 0, statusstr, attr, windowWidth)
        self._status = []

    def drawRightStatus(self, scr, prefixes, ch):
        rstat = "%s" % (keyname(ch))  # (chr(ch) if chr(ch).isprintable() else keyname(ch)
        draw_clip(scr, windowHeight-1, windowWidth-len(rstat)-2, rstat, colors[options.c_StatusLine])

    def run(self):
        global windowHeight, windowWidth
        windowHeight, windowWidth = scr.getmaxyx()
        ch = 32

        command_overrides = None
        prefixes = ''
        while True:
            if not self.sheets:
                # if no more sheets, exit
                return

            sheet = self.sheets[0]
            if sheet.nRows == 0:
                self.status('no rows')

            try:
                sheet.draw(scr)
            except Exception as e:
                self.exceptionCaught()

            self.drawLeftStatus(scr, sheet)

            self.drawRightStatus(scr, prefixes, ch)  # visible for this getch

            curses.doupdate()
            ch = scr.getch()

            self.drawRightStatus(scr, prefixes, ch)  # visible for commands that wait with getch

            if ch == curses.KEY_RESIZE:
                windowHeight, windowWidth = scr.getmaxyx()
            elif ch == curses.KEY_MOUSE:
                try:
                    devid, x, y, z, bstate = curses.getmouse()
                    sheet.cursorRowIndex = sheet.topRowIndex+y-1
                except Exception:
                    self.exceptionCaught()
            elif (prefixes, ch) in sheet.commands:
                cmdstr, helpstr = sheet.commands.get((prefixes, ch))
                try:
                    sheet.exec_command(globals(), prefixes, ch)
                    prefixes = ''
                except CommandPrefixException as e:
                    prefixes += str(e)
                except EscapeException as e:  # user aborted
                    prefixes = ''
                    self.status(keyname(e.args[0]))
                except Exception:
                    prefixes = ''
                    self.exceptionCaught()
                    self.status(cmdstr)
            else:
                prefixes = ''
                self.status('no command for key "%s" (%d) with prefixes "%s"' % (keyname(ch), ch, prefixes))

            sheet.checkCursor()

    def push(self, vs):
        if vs:
            if vs in self.sheets:
                self.sheets.remove(vs)
            self.sheets.insert(0, vs)
            return vs
# end VisiData class

### sheet layouts
#### generic list/dict/object browsing
def pushPyObjSheet(name, pyobj, src=None):
    if isinstance(pyobj, list):
        return vd().push(SheetList(name, pyobj, src=src))
    elif isinstance(pyobj, dict):
        return vd().push(SheetDict(name, pyobj))
    elif isinstance(pyobj, object):
        return vd().push(SheetObject(name, pyobj))
    else:
        status('unknown type ' + type(pyobj))

class SheetList(Sheet):
    def __init__(self, name, obj, columns=None, src=None):
        'columns is a list of strings naming attributes on the objects within the obj'
        super().__init__(name, src or obj)
        assert isinstance(obj, list)
        self.rows = obj
        if columns:
            self.columns = AttrColumns(columns)
        elif isinstance(obj[0], dict):  # list of dict
            self.columns = [Column(k, type(obj[0][k]), lambda_colname(k)) for k in obj[0].keys()]
            self.nKeys = 1
        else:
            self.columns = [Column(name)]
        self.command(Key.ENTER, 'pushPyObjSheet("%s[%s]" % (name, cursorRowIndex), cursorRow).cursorRowIndex = cursorColIndex', 'dive into this row')

class SheetDict(Sheet):
    def __init__(self, name, mapping):
        super().__init__(name, mapping)
        self.rows = list(list(x) for x in mapping.items())
        self.columns = [
            Column('key', str, lambda_col(0)),
            Column('value', anytype, lambda_col(1)) ]
        self.nKeys = 1
        self.command(Key.ENTER, 'if cursorColIndex == 1: pushPyObjSheet(name + options.SubsheetSep + cursorRow[0], cursorRow[1])', 'dive into this value')
        self.command(Key('e'), 'source[cursorRow[0]] = cursorRow[1] = editCell(1)', 'edit this value')

class SheetObject(Sheet):
    def __init__(self, name, obj):
        super().__init__(name, obj)
        self.command(Key.ENTER, 'pushPyObjSheet(name + options.SubsheetSep + cursorRow[0], cursorRow[1]() if callable(cursorRow[1]) else cursorRow[1])', 'dive into this value')
        self.command(Key('e'), 'setattr(source, cursorRow[0], editCell(1)); reload()', 'edit this value')
        self.reload()

    def reload(self):
        valfunc = lambda_col(1)
        valfunc.setter = lambda r,v,obj=self.source: setattr(obj, r[0], v)
        self.columns = [
            Column(type(self.source).__name__ + '_attr', str, lambda_col(0)),
            Column('value', anytype, valfunc) ]
        self.nKeys = 1
        self.rows = [(k, getattr(self.source, k)) for k in dir(self.source)]

#### specialized meta sheets
@functools.lru_cache()
def CommandHelp(sheet):
    vs = Sheet(sheet.name + '_help', sheet)
    vs.rows = list(sheet.commands.items())
    vs.columns = [
        Column('key', str, lambda r: r[0][0] + keyname(r[0][1])),
        Column('action', str, lambda r: r[1][1]),
        Column('global_action', str, lambda r,sheet=sheet: sheet.commands.get(('g', r[0][1]), ('', '-'))[1])
    ]
    return vs

class Sheets(SheetList):
    def __init__(self, src):
        super().__init__('sheets', vd().sheets, 'name nRows nCols cursorValue keyColNames source'.split())

        self.nKeys = 1
        self.command(Key.ENTER,    'moveListItem(vd.sheets, cursorRowIndex, 0); vd.sheets.pop(1)', 'go to this sheet')
        self.command(Key('&'), 'vd.sheets[0] = SheetJoin(selectedRows, jointype="&")', 'open inner join of selected sheets')
        self.command(Key('+'), 'vd.sheets[0] = SheetJoin(selectedRows, jointype="+")', 'open outer join of selected sheets')
        self.command(Key('*'), 'vd.sheets[0] = SheetJoin(selectedRows, jointype="*")', 'open full join of selected sheets')
        self.command(Key('~'), 'vd.sheets[0] = SheetJoin(selectedRows, jointype="~")', 'open diff join of selected sheets')
        self.rows = vd().sheets

class SheetColumns(Sheet):
    def __init__(self, srcsheet):
        super().__init__(srcsheet.name + '_columns', srcsheet)

        # on the Columns sheet, these affect the 'row' (column in the source sheet)
        self.command(Key('@'), 'cursorRow.type = date; cursorDown(+1)', 'set source column type to datetime')
        self.command(Key('#'), 'cursorRow.type = int; cursorDown(+1)', 'set source column type to integer')
        self.command(Key('$'), 'cursorRow.type = str; cursorDown(+1)', 'set source column type to string')
        self.command(Key('%'), 'cursorRow.type = float; cursorDown(+1)', 'set source column type to decimal numeric type')
        self.command(Key('~'), 'cursorRow.type = detectType(cursorRow.getValue(source.cursorRow)); cursorDown(+1)', 'autodetect type of source column using its data')
        self.command(Key('!'), 'source.toggleKeyColumn(cursorRowIndex)', 'toggle key column on source sheet')
        self.command(Key('-'), 'cursorRow.width = 0', 'hide column on source sheet')
        self.command(Key('_'), 'cursorRow.width = cursorRow.getMaxWidth(source.rows)', 'set source column width to max width of its rows')
        self.reload()

    def reload(self):
        self.rows = self.source.columns
        self.nKeys = 1
        self.columns = [
            ColumnAttr('name', str),
            ColumnAttr('width', str),
            Column('type',   str, lambda r: r.type.__name__),
            ColumnAttr('fmtstr', str),
            ColumnAttr('expr', str),
            Column('value',  anytype, lambda c,sheet=self.source: c.getValue(sheet.cursorRow)),
#            Column('nulls',  int, lambda c,sheet=sheet: c.nEmpty(sheet.rows)),

#            Column('uniques',  int, lambda c,sheet=sheet: len(set(c.values(sheet.rows))), width=0),
#            Column('mode',   anytype, lambda c: statistics.mode(c.values(sheet.rows)), width=0),
#            Column('min',    anytype, lambda c: min(c.values(sheet.rows)), width=0),
#            Column('median', anytype, lambda c: statistics.median(c.values(sheet.rows)), width=0),
#            Column('mean',   float, lambda c: statistics.mean(c.values(sheet.rows)), width=0),
#            Column('max',    anytype, lambda c: max(c.values(sheet.rows)), width=0),
#            Column('stddev', float, lambda c: statistics.stdev(c.values(sheet.rows)), width=0),
            ]

#### slicing and dicing
class SheetJoin(Sheet):
    def __init__(self, sheets, jointype='&'):
        super().__init__(jointype.join(vs.name for vs in sheets))
        self.source = sheets
        self.jointype = jointype
        self.reload()

    def reload(self):
        sheets = self.source
        # first element in joined row is the tuple of keys
        self.columns = []
        for colnum in range(sheets[0].nKeys):
            c = sheets[0].columns[colnum]
            self.columns.append(Column(c.name, c.type, lambda_subrow_wrap(lambda_col(colnum), 0)))
        self.nKeys = sheets[0].nKeys

        rowsBySheetKey = {}
        rowsByKey = {}

        for vs in sheets:
            rowsBySheetKey[vs] = {}
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                rowsBySheetKey[vs][key] = r

        for sheetnum, vs in enumerate(sheets):
            # subsequent elements are the rows from each source, in order of the source sheets
            self.columns.extend(Column(c.name, c.type, lambda_subrow_wrap(c.func, sheetnum+1), c.width) for c in vs.columns[vs.nKeys:])
            for r in vs.rows:
                key = tuple(c.getValue(r) for c in vs.keyCols)
                if key not in rowsByKey:
                    rowsByKey[key] = [key] + [rowsBySheetKey[vs2].get(key) for vs2 in sheets]  # combinedRow

        if self.jointype == '&':  # inner join  (only rows with matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if all(combinedRow))
        elif self.jointype == '+':  # outer join (all rows from first sheet)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if combinedRow[1])
        elif self.jointype == '*':  # full join (keep all rows from all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items())
        elif self.jointype == '~':  # diff join (only rows without matching key on all sheets)
            self.rows = list(combinedRow for k, combinedRow in rowsByKey.items() if not all(combinedRow))


class SheetFreqTable(Sheet):
    def __init__(self, sheet, col):
        fqcolname = '%s_%s_freq' % (sheet.name, col.name)
        super().__init__(fqcolname, sheet)

        self.origCol = col
        self.values = collections.defaultdict(list)
        for r in sheet.rows:
            self.values[str(col.getValue(r))].append(r)
        self.reload()

    def reload(self):
        self.rows = sorted(self.values.items(), key=lambda r: len(r[1]), reverse=True)  # sort by num reverse
        self.largest = len(self.rows[0][1])+1

        self.columns = [
            Column(self.origCol.name, self.origCol.type, lambda_col(0)),
            Column('num', int, lambda r: len(r[1])),
            Column('percent', float, lambda r: len(r[1])*100/sheet.source.nRows),
            Column('histogram', str, lambda r,s=self: options.ch_Histogram*int(len(r[1])*80/s.largest), width=80)
        ]
        self.nKeys = 1

        self.command(Key(' '), 'source.toggle(cursorRow[1])', 'toggle these entries')
        self.command(Key('s'), 'source.select(cursorRow[1])', 'select these entries')
        self.command(Key('u'), 'source.unselect(cursorRow[1])', 'unselect these entries')
        self.command(Key.ENTER, 'vd.push(copy(source)).rows = values[str(columns[0].getValue(cursorRow))]', 'push new sheet with only this value')


### input formats and helpers

sourceCache = {}

class Path:
    def __init__(self, fqpn):
        self.fqpn = fqpn
        self.name = os.path.split(fqpn)[-1]
        self.suffix = os.path.splitext(self.name)[1][1:]
    def read_text(self):
        return open(self.resolve(), encoding=options.encoding, errors=options.encoding_errors).read()
    def is_dir(self):
        return os.path.isdir(self.resolve())
    def iterdir(self):
        return [self.parent] + [Path(os.path.join(self.fqpn, f)) for f in os.listdir(self.resolve())]
    def stat(self):
        return os.stat(self.resolve())
    def resolve(self):
        return os.path.expandvars(os.path.expanduser(self.fqpn))
    @property
    def parent(self):
        return Path(self.fqpn + "/..")
    def __str__(self):
        return self.fqpn

def getTextContents(p):
    if not p in sourceCache:
        sourceCache[p] = p.read_text()
    return sourceCache[p]

def open_source(p):
    if isinstance(p, Path):
        if p.is_dir():
            vs = SheetDirectory(p)
        else:
            vs = globals().get('open_' + p.suffix, open_txt)(p)
    elif '://' in p:
        vs = openUrl(p)
    else:
        return open_source(Path(p))
    if vs:
        status('opened %s' % p.name)
    if isinstance(vs, Sheet):
        return vd().push(vs)

class SheetText(Sheet):
    def __init__(self, name, content, src=None):
        super().__init__(name, src)
        self.columns = [Column(name, str)]
        if isinstance(content, list):
            self.rows = content
        elif isinstance(content, str):
            self.rows = content.split('\n')
        else:
            error('unknown text type ' + str(type(content)))

class SheetDirectory(Sheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        self.columns = [Column('filename', str, lambda r: r[0].name),
                        Column('type', str, lambda r: r[0].is_dir() and '/' or r[0].suffix),
                        Column('size', int, lambda r: r[1].st_size),
                        Column('mtime', date, lambda r: r[1].st_mtime)]
        self.command(Key.ENTER, 'open_source(cursorRow[0])', 'open file')  # path, filename
        self.reload()

    def reload(self):
        self.rows = [(p, p.stat()) for p in self.source.iterdir()]  #  if not p.name.startswith('.')]

def open_txt(p):
    contents = getTextContents(p)
    if '\t' in contents[:32]:
        return open_tsv(p)  # TSV often have .txt extension
    return SheetText(p.name, contents, p)

class open_csv(Sheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        contents = getTextContents(p)

        if options.csv_dialect == 'sniff':
            headers = contents[:1024]
            dialect = csv.Sniffer().sniff(headers)
            status('sniffed csv_dialect as %s' % dialect)
        else:
            dialect = options.csv_dialect

        rdr = csv.reader(io.StringIO(contents, newline=''), dialect=dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        self.rows = [r for r in rdr]
        if options.csv_header:
            self.columns = ArrayNamedColumns(self.rows[0])
            self.rows = self.rows[1:]
        else:
            self.columns = ArrayColumns(len(self.rows[0]))

class open_tsv(Sheet):
    def __init__(self, p):
        super().__init__(p.name, p)
        lines = getTextContents(p).splitlines()

        if options.csv_header:
            self.columns = ArrayNamedColumns(lines[0].split('\t'))
            lines = lines[1:]
        else:
            self.columns = ArrayColumns(len(lines[0].split('\t')))

        self.rows = [L.split('\t') for L in lines]  # [rownum] -> [ field, ... ]

def open_json(p):
    import json
    pushPyObjSheet(p.name, json.loads(getTextContents(p)))

#### .xlsx
class open_xlsx(Sheet):
    def __init__(self, path):
        super().__init__(path.name, path)
        import openpyxl
        self.workbook = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)
        self.columns = [Column('name', str)]
        self.command(Key.ENTER, 'vd.push(getSheet(cursorRow))', 'open this sheet')

    def getSheet(self, sheetname):
        'create actual Sheet from xlsx sheet'
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        vs = Sheet('%s%s%s' % (self.source, options.SubsheetSep, sheetname), worksheet)
        vs.columns = ArrayColumns(worksheet.max_column)
        vs.rows = [ [cell.value for cell in row] for row in worksheet.iter_rows()]
        return vs

#### .hdf5
class SheetH5Obj(Sheet):
    def __init__(self, name, hobj, src):
        super().__init__(name, src)
        self.hobj = hobj
        self.reload()

    def reload(self):
        import h5py
        if isinstance(self.hobj, h5py.Group):
            self.rows = [ self.hobj[objname] for objname in self.hobj.keys() ]
            self.columns = [
                Column(self.hobj.name, str, lambda r: r.name.split('/')[-1]),
                Column('type', str, lambda r: type(r).__name__),
                Column('nItems', int, lambda r: len(r)),
            ]
            self.command(Key.ENTER, 'vd.push(SheetH5Obj(name+options.SubsheetSep+cursorRow.name, cursorRow, source))', 'open this group or dataset')
            self.command(Key('A'), 'vd.push(SheetDict(cursorRow.name + "_attrs", cursorRow.attrs))', 'open metadata sheet for this object')
        elif isinstance(self.hobj, h5py.Dataset):
            if len(self.hobj.shape) == 1:
                self.rows = self.hobj[:]  # copy
                self.columns = [Column(colname, str, lambda_colname(colname)) for colname in self.hobj.dtype.names]
            elif len(self.hobj.shape) == 2:  # matrix
                self.rows = self.hobj[:]  # copy
                self.columns = ArrayColumns(self.hobj.shape[1])
            else:
                status('too many dimensions in shape %s' % str(self.hobj.shape))
        else:
            status('unknown h5 object type %s' % type(self.hobj))

class open_hdf5(SheetH5Obj):
    def __init__(self, p):
        import h5py
        super().__init__(p.name, h5py.File(str(p), 'r'), p)
open_h5 = open_hdf5


class open_zip(Sheet):
    def __init__(self, p):
        import zipfile
        super().__init__(p.name, p)
        self.zfp = zipfile.ZipFile(p.fqpn, 'r')
        self.rows = self.zfp.infolist()
        self.columns = AttrColumns("filename file_size date_time compress_size".split())
        self.command(Ctrl.J, 'vd.push(open(cursorRow))', 'open this file')

    def open(self, zi):
        cachefn = zi.filename
        if not os.path.exists(cachefn):
            self.zfp.extract(zi)
        return open_source(cachefn)

#### databases
class SheetBlaze(Sheet):
    def __init__(self, name, data, src):
        super().__init__(name, src)
        self.columns = ArrayNamedColumns(data.fields)
        self.rows = list(data)

def openUrl(url):
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
    if m:
        return open_gspreadsheet(Path(m.group(1)))

    import blaze
    import datashape; datashape.coretypes._canonical_string_encodings.update({"utf8_unicode_ci": "U8"})
    fp = blaze.data(url)
    vs = SheetList(url, [getattr(fp, tblname) for tblname in fp.fields], url)
    vs.command(Key.ENTER, 'vd.push(SheetBlaze(cursorRow.name, cursorRow, sheet))', 'open this table')
    return vs

#### Google Sheets; requires credentials to be setup already

@functools.lru_cache()
def google_sheets():
    import httplib2
    import os

    from apiclient import discovery
    from oauth2client.file import Storage

    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    credential_path = os.path.join(credential_dir, 'sheets.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials:
        status('Credentials required')
    elif credentials.invalid:
        status('No or Invalid credentials')
    else:
        http = credentials.authorize(httplib2.Http())
        discoveryUrl = ('https://sheets.googleapis.com/$discovery/rest?version=v4')
        service = discovery.build('sheets', 'v4', http=http, discoveryServiceUrl=discoveryUrl)

        return service.spreadsheets()

def open_gspreadsheet(p):
    sheets = google_sheets()
    sheet_md = sheets.get(spreadsheetId=p.name).execute()
    vs = Sheet(sheet_md['properties']['title'], p)
    vs.columns = [Column('title', lambda_eval('properties.title')),
                  Column('rowCount', lambda_eval('properties.gridProperties.rowCount')),
                  Column('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
    vs.rows = sheet_md['sheets']
    vs.command(Key.ENTER, 'cursorRow')
    return vs

def open_gsheet(p):
    sheets = google_sheets()
    sheet = sheets.values().get(spreadsheetId=p.name).execute()
    pushPyObjSheet(p.name, sheet, p)
#    vs = Sheet(sheet_md['properties']['title'], p)
#    vs.columns = [Column('title', lambda_eval('properties.title')),
#                  Column('rowCount', lambda_eval('properties.gridProperties.rowCount')),
#                  Column('columnCount', lambda_eval('properties.gridProperties.columnCount'))]
#    vs.rows = sheet_md['sheets']
#    vs.command(Key.ENTER, 'cursorRow')
#    return vs

#### external addons
def open_py(p):
    contents = getTextContents(p)
    exec(contents, globals())
    status('executed %s' % p)

### Sheet savers

def saveSheet(sheet, fn):
    if options.readonly:
        status('readonly mode')
        return
    basename, ext = os.path.splitext(fn)
    funcname = 'save_' + ext[1:]
    globals().get(funcname, save_tsv)(sheet, fn)
    status('saved to ' + fn)

def save_tsv(sheet, fn):
    with open(fn, 'w', encoding=options.encoding, errors=options.encoding_errors) as fp:
        colhdr = '\t'.join(col.name for col in sheet.visibleCols) + '\n'
        if colhdr.strip():  # is anything but whitespace
            fp.write(colhdr)
        for r in sheet.rows:
            fp.write('\t'.join(col.getDisplayValue(r) for col in sheet.visibleCols) + '\n')

def save_csv(sheet, fn):
    with open(fn, 'w', newline='', encoding=options.encoding, errors=options.encoding_errors) as fp:
        cw = csv.writer(fp, dialect=options.csv_dialect, delimiter=options.csv_delimiter, quotechar=options.csv_quotechar)
        colnames = [col.name for col in sheet.visibleCols]
        if ''.join(colnames):
            cw.writerow(colnames)
        for r in sheet.rows:
            cw.writerow([col.getDisplayValue(r) for col in sheet.visibleCols])

### curses, options, init

def inputLine(prompt, value=''):
    'add a prompt to the bottom of the screen and get a line of input from the user'
    scr.addstr(windowHeight-1, 0, prompt)
    return vd().editText(windowHeight-1, len(prompt), windowWidth-len(prompt)-8, value=value, attr=colors[options.c_EditCell], unprintablechar=options.ch_Unprintable)


def run():
    os.putenv('ESCDELAY', '25')  # reduce ESC timeout to 25ms. http://en.chys.info/2009/09/esdelay-ncurses/
    ret = wrapper(curses_main)
    if ret:
        print(ret)

def curses_main(_scr):
    global scr
    scr = _scr

    try:
        return vd().run()
    except Exception as e:
        if options.debug:
            raise
        return '%s: %s' % (type(e).__name__, str(e))


setup_sheet_commands()  # on module import
